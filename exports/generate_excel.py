import re
import os
import csv
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 1. Parse products from products-data.ts
with open('d:/nuvaanah/lib/products-data.ts', 'r', encoding='utf-8') as f:
    lines = f.readlines()

products = []
current_product = None
in_product = False

for line in lines:
    if line.startswith('  {'):
        current_product = {}
        in_product = True
        continue
    
    if line.startswith('  }') or line.startswith('  },'):
        if in_product and current_product and 'name' in current_product:
            products.append(current_product)
        current_product = None
        in_product = False
        continue
    
    if in_product and current_product is not None:
        match = re.match(r'^    ([a-zA-Z0-9_]+)\s*:\s*(.*),?$', line)
        if match:
            key = match.group(1)
            val_str = match.group(2).strip()
            
            if val_str.endswith(','):
                val_str = val_str[:-1].strip()
                
            if val_str.startswith("'") and val_str.endswith("'"):
                val = val_str[1:-1]
            elif val_str.startswith('"') and val_str.endswith('"'):
                val = val_str[1:-1]
            elif val_str.isdigit():
                val = int(val_str)
            elif val_str == 'true':
                val = True
            elif val_str == 'false':
                val = False
            else:
                val = val_str
                
            current_product[key] = val

# 2. Formulate pricing rules and clean up descriptions
# Old Price is a markup over the new price (which is the current catalog price).
def calculate_prices(p_name, current_price):
    try:
        price_num = int(current_price)
    except:
        price_num = 0

    if price_num == 0:
        return 0, 0

    # Pricing markup rule
    if price_num < 2000:
        old_price = int(price_num * 1.25)
        # round to nearest 50
        old_price = ((old_price + 25) // 50) * 50
    elif price_num < 10000:
        old_price = int(price_num * 1.20)
        # round to nearest 100
        old_price = ((old_price + 50) // 100) * 100
    else:
        old_price = int(price_num * 1.25)
        # round to nearest 1000
        old_price = ((old_price + 500) // 1000) * 1000

    return old_price, price_num

catalog_data = []
for p in products:
    name = p.get('name', '')
    desc = p.get('description', '')
    price_raw = p.get('price', 0)
    
    # Clean name
    name_clean = name.split('|')[0].strip()
    
    # Clean description (remove escape sequences)
    desc_clean = desc.replace('\\n', '\n').replace('\\t', '\t').replace("\\'", "'").replace('', '—')
    
    old_price, new_price = calculate_prices(name_clean, price_raw)
    
    catalog_data.append({
        'name': name_clean,
        'description': desc_clean,
        'price': new_price,
        'old_price': old_price,
        'new_price': new_price
    })

# 3. Generate CSV version
csv_path = 'd:/nuvaanah/exports/products_catalog.csv'
with open(csv_path, 'w', encoding='utf-8', newline='') as f:
    writer = csv.writer(f)
    writer.writerow(['Product Name', 'Description', 'Price', 'Old Price', 'New Price'])
    for p in catalog_data:
        writer.writerow([p['name'], p['description'], p['price'], p['old_price'], p['new_price']])

print(f"CSV catalog successfully generated at: {csv_path}")

# 4. Generate beautifully styled Excel version
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Products Catalog"

# Enable grid lines
ws.views.sheetView[0].showGridLines = True

# Add title block
title_font = Font(name="Segoe UI", size=16, bold=True, color="0D5C75")
subtitle_font = Font(name="Segoe UI", size=10, italic=True, color="555555")

ws["A1"] = "Nuvaanah Product Catalog"
ws["A1"].font = title_font
ws["A2"] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
ws["A2"].font = subtitle_font

# Empty spacer row
ws.row_dimensions[1].height = 25
ws.row_dimensions[2].height = 18
ws.row_dimensions[3].height = 15

# Table Headers
headers = ['Product Name', 'Description', 'Price', 'Old Price', 'New Price']
header_fill = PatternFill(start_color="0D5C75", end_color="0D5C75", fill_type="solid")
header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color='E0E0E0'),
    right=Side(style='thin', color='E0E0E0'),
    top=Side(style='thin', color='E0E0E0'),
    bottom=Side(style='thin', color='E0E0E0')
)

ws.row_dimensions[4].height = 28
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align
    cell.border = thin_border

# Write Data Rows
data_font = Font(name="Segoe UI", size=11, color="333333")
price_font = Font(name="Segoe UI", size=11, bold=True, color="000000")
align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
align_right = Alignment(horizontal="right", vertical="top")

row_num = 5
for p in catalog_data:
    ws.row_dimensions[row_num].height = 35  # tall enough for wrapped description
    
    # Product Name
    c_name = ws.cell(row=row_num, column=1)
    c_name.value = p['name']
    c_name.font = data_font
    c_name.alignment = align_left
    c_name.border = thin_border
    
    # Description
    c_desc = ws.cell(row=row_num, column=2)
    c_desc.value = p['description']
    c_desc.font = data_font
    c_desc.alignment = align_left
    c_desc.border = thin_border
    
    # Price
    c_price = ws.cell(row=row_num, column=3)
    c_price.value = p['price']
    c_price.font = price_font
    c_price.alignment = align_right
    c_price.number_format = '₹#,##0'
    c_price.border = thin_border
    
    # Old Price
    c_old = ws.cell(row=row_num, column=4)
    c_old.value = p['old_price']
    c_old.font = data_font
    c_old.alignment = align_right
    c_old.number_format = '₹#,##0'
    c_old.border = thin_border
    
    # New Price
    c_new = ws.cell(row=row_num, column=5)
    c_new.value = p['new_price']
    c_new.font = price_font
    c_new.alignment = align_right
    c_new.number_format = '₹#,##0'
    c_new.border = thin_border
    
    row_num += 1

# Set Column Widths
ws.column_dimensions['A'].width = 25  # Name
ws.column_dimensions['B'].width = 65  # Description
ws.column_dimensions['C'].width = 15  # Price
ws.column_dimensions['D'].width = 15  # Old Price
ws.column_dimensions['E'].width = 15  # New Price

xlsx_path = 'd:/nuvaanah/exports/products_catalog.xlsx'
wb.save(xlsx_path)
print(f"Styled Excel catalog successfully generated at: {xlsx_path}")
